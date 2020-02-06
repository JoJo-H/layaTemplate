from openpyxl import Workbook
from openpyxl import load_workbook
import os
from loadconfig import loadPath
import sys
import time

def parse():
	ts = loadPath()
	# excel所在的目录
	ep = ts[ 0 ]

	excelList = []
	# 获得所有excel文件
	for file in os.listdir(ep):
		dir = os.path.join(ep, file)
		if os.path.isfile(dir):
			if not file.startswith('~') and file.endswith(".xlsx"):
				excelList.append(dir)
	titles = []
	clientDatas = []
	# 解析excel
	for dir in excelList:
		print(dir)
		wb = load_workbook(dir, read_only=True, data_only=True)
		sheetNames = wb.get_sheet_names()
		for sheetName in sheetNames:
			sheet = wb.get_sheet_by_name(sheetName)  #获取工作表
			# 只解析tb开头的
			if sheet.title.find('tb_') == -1:
				continue
			cdata, serverTitle = parseSheet(sheet, ts)
			if cdata:
				clientDatas.append(cdata)
			if serverTitle:
				titles.append(serverTitle)
	

	# 生成客户端
	dir = os.path.join(ts[ 1 ], 'tb.txt')
	with open(dir, 'wb') as fp:
		writeBinary(fp, '\r'.join(clientDatas))
	print('一共生成了：',len(excelList),'表')
	'''
	# 生成init.lua
	dir = os.path.join(ts[ 2 ], 'init.lua')
	with open(dir, 'wb') as fp:
		for title in titles:
			writeBinary(fp, 'require("template.tables.{0}")\r\n'.format(title))
	'''
	
	

def parseSheet(sheet, ts):
	max_row = sheet.max_row
	max_column = sheet.max_column
	
	print("parse", sheet.title)
	data_name = sheet.title
	prefix = data_name[3:]
	#print(sheet.title, sheet.max_row, sheet.max_column)
	
	# 解析头
	flags = []
	for i in range(max_column):
		value = sheet.cell(row = 1, column = i+1).value
		flags.append(value)
		
	types = []
	for i in range(max_column):
		value = sheet.cell(row = 2, column = i+1).value
		types.append(value)
	
	vars = []
	for i in range(max_column):
		value = sheet.cell(row = 3, column = i+1).value
		vars.append(value)
	
	comments = []
	for i in range(max_column):
		value = sheet.cell(row = 4, column = i+1).value
		comments.append(value)
	
	#print(sheet.rows)
	#ta = time.time()

	# 解析内容
	clientData = []
	serverData = []
	consntData = []
	pythonData = []
	#ta = time.time()
	iters = sheet.iter_rows(min_row = 5)
	row = 4
	for rows in iters:
		row = row + 1
		
		# 如果第一个就没数据 表示下面就是空的 不进行解析
		if rows[ 0 ].value is None:
			break
		
		clientInfo = []
		serverInfo = []
		pythonInfo = []
		ct = False
		col = 0
		id_value = None
		for cell in rows:
			j = col
			col = col + 1
			value = cell.value
			# 是否服务端需要
			if flags[ j ] and flags[ j ].find('S') > -1:
				#try:
				_value = parseServerType(types[ j ], value)
				#except Exception as err:
				#	print(err, " in row:", row, " in col:", col, " value = ", repr(value))
				#	sys.exit(0)
				if j == 0:
					serverInfo.append(_value)
					id_value = _value
				serverInfo.append('{0} = {1}'.format(vars[ j ], _value))
			
			if flags[ j ] and flags[ j ].find('G') > -1:
				#try:
				_value = parseServerType(types[ j ], value, True)
				#except Exception as err:
				#	print(err, " in row:", row, " in col:", col, " value = ", repr(value))
				#	sys.exit(0)
				if j == 0:
					pythonInfo.append(_value)
					id_value = _value
				pythonInfo.append('"{0}" : {1}'.format(vars[ j ], _value))

			# 是否客户端需要
			if flags[ j ] and flags[ j ].find('C') > -1:
				_value = parseClientType(types[ j ], value)
				clientInfo.append(_value)
				
			if flags[ j ] and flags[ j ].find('P') > -1 and not ct:
				ct = True
				str_value = str(value)
				str2_value = str(_value)				
				vv = "_M." + prefix.upper() + "_" + str_value.upper() + " = " + str(id_value)
				consntData.append(vv)
			
			if flags[ j ] and flags[ j ].find('H') > -1:
				if len(consntData) > 0:
					str_value = '	-- ' + str(value)
					consntData[-1] = consntData[-1] + str_value
				
		if len(clientInfo) > 1:
			clientData.append(clientInfo)

		if len(serverInfo) > 1:
			serverData.append(serverInfo)
			
		if len(pythonInfo) > 1:
			pythonData.append(pythonInfo)
	#tb = time.time()
	#print("load waste", (tb-ta))
	
	# printArray(clientData)
	# printArray(serverData)
	
	# 写服务端数据
	serverTitle = None
	if len(serverData) > 0:
		serverTitle = data_name
		dir = os.path.join(ts[ 2 ], data_name + '.lua')
		with open(dir, 'wb') as fp:
			writeBinary(fp, "--[[\r\n")
			writeBinary(fp, "	文件是通过excel自动生成的\r\n")
			writeBinary(fp, "--]]\r\n")
			writeBinary(fp, "\r\n")
			writeBinary(fp, "\r\n")
			writeBinary(fp, "local " + sheet.title + " = {\r\n")
			# 字段说明
			for i in range(len(flags)):
				flag = flags[ i ]
				if flag and flag.find('S') > -1:
					writeBinary(fp, "	--  {0}:{1} {2}\r\n".format(vars[ i ], types[ i ], comments[ i ]))
			
			for info in serverData:
				value = '{' + ', '.join(info[1:]) + '}'
				writeBinary(fp, "	[{0}] = {1},\r\n".format(info[ 0 ], value))

			writeBinary(fp, "}\r\n")
			writeBinary(fp, "\r\nreturn " + sheet.title)
			
	serverTitle = None
	if len(pythonData) > 0:
		serverTitle = data_name
		dir = os.path.join(ts[ 3 ], data_name + '.py')
		with open(dir, 'wb') as fp:
			writeBinary(fp, "'''\r\n")
			writeBinary(fp, "	文件是通过excel自动生成的\r\n")
			writeBinary(fp, "'''\r\n")
			writeBinary(fp, "\r\n")
			writeBinary(fp, "\r\n")
			writeBinary(fp, sheet.title + " = {\r\n")
			# 字段说明
			for i in range(len(flags)):
				flag = flags[ i ]
				if flag and flag.find('S') > -1:
					writeBinary(fp, "	#  {0}:{1} {2}\r\n".format(vars[ i ], types[ i ], comments[ i ]))
			
			for info in pythonData:
				value = '{' + ', '.join(info[1:]) + '}'
				writeBinary(fp, "	{0} : {1},\r\n".format(info[ 0 ], value))

			writeBinary(fp, "}\r\n")
			#writeBinary(fp, "\r\nreturn " + sheet.title)
			
			
	if len(consntData) > 0:
		dir = os.path.join(ts[ 2 ], 'ct_' + prefix + '.lua')
		with open(dir, 'wb') as fp:
			writeBinary(fp, "local _M = {}\r\n\r\n")
			writeBinary(fp, "\r\n".join(consntData))
			writeBinary(fp, "\r\n\r\nreturn _M")
			
	
	string = None
	# 客户端的数据返回写成一个字符串
	if len(clientData) > 0:
		cvars = []
		ctypes = []
		for i in range(len(flags)):
			flag = flags[ i ]
			if flag and flag.find('C') > -1:
				cvars .append(vars [ i ])
				ctypes.append(types[ i ])

		string = data_name + '\r'
		string += ','.join(cvars) + '\r'
		string += ','.join(ctypes) + '\r'
		linerdata = []
		for cdata in clientData:
			linerdata.extend(cdata)
		string += '\01'.join(linerdata)
		
		#print(string)
		
	print(sheet.title, "...OK")
	
	return string, serverTitle
	
def writeBinary(fp, string):
	fp.write(string.encode(encoding = "utf-8"))
	
def parseServerType(type, value, isp = False):
	if type == 'int':
		if not value:
			return 0
		return int(value)
	elif type == 'bool':
		if not value:
			return "false"
		return str(value)
	elif type == 'float':
		if not value:
			return "0.00"
		return value
	elif type.startswith("array"):
		if not value and value != 0:
			return isp and "(,)" or "{}"
		return isp and "(" + str(value) + ")" or "{" + str(value) + "}"
	elif type == 'string':
		if not value and value != 0:
			return "\"\""
		return "\"" + str(value) + "\""
	else:
		raise Exception("无法解析的数据类型:" + type)
	return None
	
def parseClientType(type, value):
	if type == 'int':
		if value is None:
			return '0'
		return str(value)
	elif type == 'bool':
		if not value:
			return "false"
		return str(value)
	elif type == 'float':
		if not value:
			return "0.00"
		return value
	elif type.startswith("array"):
		if value is None:
			return ""
			
		return str(value)
	elif type == 'string':
		if value is None:
			return ""
		return str(value)
	else:
		raise Exception("无法解析的数据类型:" + type)
	return None
	
def printArray(list):
	for element in list:
		print(element)
	if len(list) > 0:
		print()

if __name__ == "__main__":
	parse()